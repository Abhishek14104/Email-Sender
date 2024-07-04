import openpyxl

workbook = openpyxl.load_workbook('list.xlsx')

sheet = workbook.active

# cellValue = sheet['A1'].value
# print(cellValue)
EMAIL_ADDRESS = []
userName = []

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
    EMAIL_ADDRESS.append(row[0])
    userName.append(row[1])
    
# print(EMAIL_ADDRESS[1])
# print(userName)

for i in range(len(EMAIL_ADDRESS)):
    print(EMAIL_ADDRESS[i])
    print(userName[i])
    
workbook.close()