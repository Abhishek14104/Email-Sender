import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl

EMAIL_ADDRESS = sys.argv[1]
passwordList = sys.argv[2:]

EMAIL_PASSWORD = " ".join(passwordList)

if EMAIL_ADDRESS is None or EMAIL_PASSWORD is None:
    raise ValueError("Please enter your Email then your Password")

# Esse STMP server start hota h
s = smtplib.SMTP('smtp.gmail.com', 587)
s.starttls()

# Apni provided Id Password se login ho gaye
s.login(EMAIL_ADDRESS, EMAIL_PASSWORD)

# Ab yaha se apna xl sheet ko activate and read karana shuru hua h
workbook = openpyxl.load_workbook('list.xlsx')
sheet = workbook.active

try:
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        recipientEmail = row[0]
        recipientUserName = row[1]
        
        # Check if the recipientEmail or recipientUserName is not provided for any row
        if not recipientEmail or not recipientUserName:
            raise ValueError(f"Invalid data in Excel row: Email={recipientEmail}, Username={recipientUserName}")
        
        # Message Body bana shuru ho gaya
        msg = MIMEMultipart()
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = recipientEmail
        msg['Subject'] = "Automated Email"
        body = f'This is my first automated email sent to {recipientUserName} using Python.'

        msg.attach(MIMEText(body, 'plain'))

        s.sendmail(EMAIL_ADDRESS, recipientEmail, msg.as_string())
    
        print(f"Email sent to {recipientEmail} named as {recipientUserName}")
        
except ValueError as ve:
    print(f"Error: {ve}")

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    workbook.close()
    s.quit()
